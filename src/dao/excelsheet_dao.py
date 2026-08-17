from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Iterator, Any
from pathlib import Path
from datetime import datetime


class ExcelSheetReadDao:
    def __init__(self,
                 excel_file: str,
                 sheet_index: int,
                 start_row_index : int,
                 col_index: tuple[int, ...]) -> None:

        self._col_index = col_index

        self._it_row, self._wb = self._read(excel_file, sheet_index, start_row_index, min(col_index), max(col_index))


    def _read(self, excel_file: str,
              sheet_index: int,
              start_row_index: int,
              min_col_index: int,
              max_col_index: int):

        wb = load_workbook(excel_file)
        ws = wb.worksheets[sheet_index]

        return ws.iter_rows(min_row = start_row_index + 1,
                                      min_col = min_col_index + 1,
                                      max_col = max_col_index + 1,
                                      values_only = True), wb


    def close(self) -> None:
        self._wb.close()


    def __iter__(self) -> Iterator[list[Any]]:
        return self


    def __next__(self) -> list[Any | str | int | float | datetime]:
        row_lst = next(self._it_row)

        res = list()

        for i in self._col_index:
            res.append(row_lst[self._col_index[i]])

        return res


class _ExcelSheetWriteDao:

    def __init__(self, excel_file: str, wb: Workbook, ws: Worksheet) -> None:
        self._excel_file = excel_file
        self._wb = wb
        self._ws = ws


    def append(self, data: list[Any]) -> None:
        self._ws.append(data)


    def save(self) -> None:
        self._wb.save(self._excel_file)


    def close(self) -> None:
        self._wb.close()



class ExcelSheetCreateDao(_ExcelSheetWriteDao):
    def __init__(self,
                 excel_file: str,
                 new_name: str,
                 index: int | None = None) -> None:

        self._excel_file = excel_file

        self._wb, self._ws = self._create(excel_file, new_name, index)


    def _create(self,
                excel_file: str,
                new_name: str,
                index: int | None = None) -> tuple[Workbook, Worksheet]:

        if not Path(excel_file).exists():
            wb = Workbook()
            wb.remove(wb.worksheets[0])

            ws = wb.create_sheet(new_name)
            return wb, ws

        else:
            wb = load_workbook(excel_file)
            ws = wb.create_sheet(new_name, index)
            return wb, ws



class ExcelSheetAppendDao(_ExcelSheetWriteDao):
    def __init__(self,
                 excel_file: str,
                 index: int) -> None:

        self._excel_file = excel_file
        self._wb, self._ws = self._read(index)


    def _read(self, index: int) -> tuple[Workbook, Worksheet]:
        wb = load_workbook(self._excel_file)
        ws = wb.worksheets[index]
        return wb, ws
