from openpyxl import Workbook

class Bill:
    def __init__(self, category, wechat=None, alipay=None):
        self._init_xl(category)

        if wechat:
            self._add_xl(wechat.wb, '分账单-微信')
        
        if alipay:
            self._add_xl(alipay.wb, '分帐单-支付宝')


    def classify(self, classifier, batch):
        for start in range(2, self.wb['总账单'].max_row+1, batch):
            end = min(start + batch - 1, self.wb['总账单'].max_row)
            
            group = list()

            for row in self.wb['总账单'].iter_rows(start, end, values_only = True):
                group.append(
                    {
                        '交易时间':row[0],
                        '交易对方':row[1],
                        '交易商品':row[2],
                        '收支类型':row[3],
                        '金额':row[4]
                    }
                )

            res = classifier(group)

            for idx in range(len(res)):
                row_value = next(self.wb['总账单'].iter_rows(start + idx, start + idx, values_only = True))
                self.wb[res[idx]].append(row_value)


    def save(self, path):
        self.wb.save(path)


    def _init_xl(self, category):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

        ws = self.wb.create_sheet('总账单')
        ws.append(['交易时间', '交易对方', '交易商品', '收支类型', '金额'])

        for c in category:
            ws = self.wb.create_sheet(c)
            ws.append(['交易时间', '交易对方', '交易商品', '收支类型', '金额'])


    def _add_xl(self, from_wb, name):
        ws_cp = self.wb.create_sheet(name, 1)
        for row in from_wb.worksheets[0].iter_rows(1,values_only = True):
            ws_cp.append(row)

        ws_ap = self.wb['总账单']
        for row in from_wb.worksheets[0].iter_rows(2,values_only = True):
            ws_ap.append(row)
