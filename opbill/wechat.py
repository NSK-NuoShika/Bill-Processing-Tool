from openpyxl import load_workbook

class Wechat:
    def __init__(self, path):
        self.load(path)
    
    def load(self, path):
        self.wb = load_workbook(path)

    def hand(self):
        self._del_redundant()
        self._standard()

    def save(self, path):
        self.wb.save(path)

    def _del_redundant(self):
        ws = self.wb.worksheets[0]
        
        for i in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(i))

        ws.delete_rows(1, 17)
        ws.delete_cols(9, 3)
        ws.delete_cols(8, 1)
        ws.delete_cols(7, 1)
        ws.delete_cols(2, 1)

    def _standard(self):
        ws = self.wb.worksheets[0]

        ws['C1'].value = '交易商品'
        ws['D1'].value = '收支类型'
        ws['E1'].value = '金额'

        for cell in next(ws.iter_cols(4,4,2)):
            if cell.value == '收入':
                cell.value = '收'
            elif cell.value == '支出':
                cell.value = '支'
            elif cell.value == '/':
                cell.value = '中'
