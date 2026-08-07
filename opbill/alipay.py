from openpyxl import load_workbook

class Alipay:
    def __init__(self, path):
        self.load(path)

    def load(self, path):
        self.wb = load_workbook(path)
    
    def hand(self):
        self._del_redundant()
        self._standerd()

    def save(self, path):
        self.wb.save(path)
    
    def _del_redundant(self):
        ws = self.wb.worksheets[0]

        for i in ws.merged_cells.ranges:
            ws.unmerge_cells(str(i))

        ws.delete_rows(1, 23)
        ws.delete_cols(8, 5)
        ws.delete_cols(4, 1)
        ws.delete_cols(2, 1)


    def _standerd(self):
        ws = self.wb.worksheets[0]

        ws['C1'] = '交易商品'
        ws['D1'] = '收支类型'

        for cell in next(ws.iter_cols(4, 4, 2)):
            if cell.value == '支出':
                cell.value = '支'
            elif cell.value == '收入':
                cell.value = '收'
            elif cell.value == '不计收支':
                cell.value = '中'