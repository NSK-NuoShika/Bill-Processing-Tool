from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelMergeDao:

    def __init__(self, to_excel_path):
        self._to_excel_path = to_excel_path
        self._wb = self._create()


    def merge(self, from_excel_path: str):
        from_wb  = load_workbook(from_excel_path)

        for from_ws in from_wb.worksheets:
            to_ws = self._wb.create_sheet(from_ws.title)
            self._copy_sheet(from_ws, to_ws)

        from_wb.close()


    def save(self) -> None:
        self._wb.save(self._to_excel_path)
        self._wb.close()


    def _create(self) -> Workbook:
        if not Path(self._to_excel_path).exists():
            wb = Workbook()
            wb.remove(wb.worksheets[0])
        else:
            wb = load_workbook(self._to_excel_path)

        return wb


    def _copy_sheet(self, from_ws: Worksheet, to_ws: Worksheet) -> None:
        for row in from_ws.iter_rows():
            for cell in row:
                to_ws[cell.coordinate].value = cell.value

        for m in from_ws.merged_cells.ranges:
            to_ws.merge_cells(str(m))
